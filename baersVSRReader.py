
from openpyxl import Workbook

import re

baersRegex = re.compile(r'''(.{4})\s(\d{3}.\d{5}.\d{1})\s(.{3})\s(.{21})\s(.{20})''')

ppr = ''

#read in ifr text file
f = open('vsr.txt', 'r')
for line in f:
    ppr += line
f.close()


lines = []

lines = baersRegex.findall(ppr)

wb = Workbook()

ws = wb.active

ws['A1'] = 'VEND'
ws['B1'] = 'SKU'
ws['C1'] = 'DROP'
ws['D1'] = 'VSN'
ws['E1'] = 'DESC'
#ws['F1'] = 'SIZE'
#ws['G1'] = 'FINISH'
#ws['E1'] = 'STYLE'
#ws['F1'] = 'RETAIL'
#ws['G1'] = 'ADV'



ws['C5'] = "you didn't copy amything"



rowcnt = 2

for line in range(len(lines)):
    VEND, SKU, DROP, VSN, DESC = lines[line]

    #print(skus[item])
    #if str(vend) in carey_codes:

    print(VEND)

    SKU = SKU[0:3] + SKU[4:9] + SKU[10:11]


    ws['A' + str(rowcnt)] = str(VEND)
    ws['B' + str(rowcnt)] = str(SKU)
    ws['C' + str(rowcnt)] = str(DROP)
    ws['D' + str(rowcnt)] = str(VSN)
    ws['E' + str(rowcnt)] = str(DESC)
    #ws['F' + str(rowcnt)] = str(SIZE)
    #ws['G' + str(rowcnt)] = str(FINISH)
    #ws['H' + str(rowcnt)] = str(STYLE)
    #ws['I' + str(rowcnt)] = str(RETAIL)
    #ws['J' + str(rowcnt)] = str(ADV)


    rowcnt += 1


wb.save("VSR.xlsx")
